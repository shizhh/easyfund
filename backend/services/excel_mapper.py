"""Generic Excel-to-model parser driven by ImportMappingSchema."""

import io
import logging
import re
import uuid

import openpyxl

from backend.schemas import ImportMappingSchema

logger = logging.getLogger(__name__)


def _uid() -> str:
    return uuid.uuid4().hex[:8]


# ── Transform Application ────────────────────────────────────

def apply_transform(raw_value, transform: str | None, transform_group: int | None = None):
    """Apply a transform rule to a raw cell value."""
    if transform is None or raw_value is None:
        return raw_value

    raw = str(raw_value)

    if transform.startswith("regex:"):
        pattern = transform[6:]
        m = re.search(pattern, raw)
        if m:
            group = transform_group or 1
            try:
                return m.group(group)
            except IndexError:
                return m.group(0)
        return None

    if transform == "int":
        try:
            return int(float(raw))
        except (ValueError, TypeError):
            return 0

    if transform == "float":
        try:
            return float(raw)
        except (ValueError, TypeError):
            return 0.0

    if transform.startswith("date:"):
        from datetime import datetime
        fmt = transform[5:]
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            return raw

    if transform == "currency_code":
        m = re.search(r'[A-Z]{3}', raw)
        return m.group(0) if m else "CNY"

    return raw_value


# ── Core Parser ───────────────────────────────────────────────

def parse_workbook_with_mapping(content: bytes, mapping: ImportMappingSchema) -> dict:
    """Parse an Excel workbook using the provided mapping schema.

    Returns dict with keys: accounts, holdings, transactions, deposits, rates
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    accounts = []
    holdings = []
    transactions = []
    deposits = []
    rates_data = {}  # currency -> rate_to_cny

    for sheet_mapping in mapping.sheets:
        sheet_name = sheet_mapping.sheet_name
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found in workbook, skipping", sheet_name)
            continue

        ws = wb[sheet_name]
        header_row = sheet_mapping.header_row
        data_start_row = sheet_mapping.data_start_row

        # Build column mapping lookup: column_index -> ColumnMapping
        col_map = {}
        for cm in sheet_mapping.column_mappings:
            col_map[cm.column_index] = cm

        # Build category rule lookup
        category_lookup = {}
        for cr in sheet_mapping.category_rules:
            category_lookup[cr.raw_value] = cr.category

        # Process cell references (standalone values like exchange rates)
        for cell_ref in sheet_mapping.cell_references:
            try:
                val = ws[cell_ref.cell_ref].value
                if val is not None:
                    val = apply_transform(val, None, None)
                    try:
                        val = float(val)
                    except (ValueError, TypeError):
                        pass

                    if cell_ref.target_type == "exchange_rate":
                        currency = cell_ref.currency or "UNKNOWN"
                        rates_data[currency] = float(val)
            except Exception as e:
                logger.warning("Failed to read cell %s: %s", cell_ref.cell_ref, e)

        # Process data rows
        # Track current section if section_detection is present
        current_section = None
        section_detection = sheet_mapping.section_detection

        for row in ws.iter_rows(min_row=data_start_row + 1, max_row=ws.max_row, values_only=False):
            # Build row values list (0-indexed)
            row_values = {}
            for cell in row:
                row_values[cell.column - 1] = cell.value  # openpyxl is 1-indexed

            # Section detection: check if this row starts a new section
            if section_detection:
                marker_col = section_detection.section_marker_column
                marker_val = row_values.get(marker_col)
                if marker_val is not None and str(marker_val).strip():
                    current_section = str(marker_val).strip()

            # Apply skip conditions
            skip = False
            for cond in sheet_mapping.skip_conditions:
                # Simple condition format: "field_name == 'value'"
                m = re.match(r"(\w+)\s*==\s*['\"](.+)['\"]", cond)
                if m:
                    field_name, expected = m.group(1), m.group(2)
                    # Find the column that maps to this field
                    for col_idx, cm in col_map.items():
                        if cm.target_field == field_name:
                            actual_val = str(row_values.get(col_idx, "") or "")
                            if actual_val == expected:
                                skip = True
                                break
            if skip:
                continue

            # Extract fields from row using column mappings
            record = {}
            for col_idx, cm in col_map.items():
                if cm.target_field == "skip":
                    continue
                raw_val = row_values.get(col_idx)
                val = apply_transform(raw_val, cm.transform, cm.transform_group)
                if val is not None:
                    record[cm.target_field] = val

            # Apply category rules
            if "category_raw" in record:
                raw_cat = str(record.pop("category_raw"))
                record["category"] = category_lookup.get(raw_cat, "cash")

            # Determine target types for this row
            for target_type in sheet_mapping.target_types:
                if target_type == "account":
                    if "name" in record and record["name"]:
                        acct_id = _uid()
                        acct = {
                            "id": acct_id,
                            "name": str(record.get("name", "")),
                            "category": record.get("category", "cash"),
                            "institution": str(record.get("institution", "")),
                            "balances": record.get("balances", [{"currency": "CNY", "amount": float(record.get("amount_cny", record.get("amount", 0)) or 0)}]),
                            "notes": str(record.get("notes", "")),
                        }
                        accounts.append(acct)

                        # Check if this account also has holding info
                        if target_type == "account" and "holding_ticker" in record:
                            holdings.append({
                                "id": _uid(),
                                "account_id": acct_id,
                                "ticker": str(record.get("holding_ticker", "")),
                                "name": str(record.get("holding_name", "")),
                                "shares": int(float(record.get("holding_shares", 0) or 0)),
                                "avg_cost_price": float(record.get("holding_avg_cost", 0) or 0),
                                "current_price": float(record.get("holding_current_price", 0) or 0),
                                "currency": str(record.get("holding_currency", "USD")),
                                "vested_shares": int(float(record.get("holding_shares", 0) or 0)),
                                "unvested_shares": 0,
                                "vesting_schedule": [],
                            })

                elif target_type == "holding":
                    if "ticker" in record and record["ticker"]:
                        # Find or create account for this holding
                        acct_id = record.get("account_id", "")
                        if not acct_id:
                            # Try to match by account name
                            acct_name = record.get("account_name", current_section or "")
                            found = next((a for a in accounts if a["name"] == acct_name), None)
                            if found:
                                acct_id = found["id"]
                            else:
                                acct_id = _uid()
                                accounts.append({
                                    "id": acct_id,
                                    "name": acct_name,
                                    "category": "investment",
                                    "institution": "",
                                    "balances": [{"currency": "CNY", "amount": 0}],
                                    "notes": "",
                                })
                        holdings.append({
                            "id": _uid(),
                            "account_id": acct_id,
                            "ticker": str(record.get("ticker", "")),
                            "name": str(record.get("name", "")),
                            "shares": int(float(record.get("shares", 0) or 0)),
                            "avg_cost_price": float(record.get("avg_cost_price", 0) or 0),
                            "current_price": float(record.get("current_price", 0) or 0),
                            "currency": str(record.get("currency", "USD")),
                            "vested_shares": int(float(record.get("vested_shares", record.get("shares", 0)) or 0)),
                            "unvested_shares": int(float(record.get("unvested_shares", 0) or 0)),
                            "vesting_schedule": record.get("vesting_schedule", []),
                        })

                elif target_type == "transaction":
                    if "amount" in record and record["amount"] is not None:
                        # Resolve account_id
                        acct_id = record.get("account_id", "")
                        if not acct_id:
                            acct_name = record.get("account_name", current_section or "")
                            found = next((a for a in accounts if a["name"] == acct_name), None)
                            if found:
                                acct_id = found["id"]
                            else:
                                # Map common account names
                                acct_map = {"富途": "futu", "支付宝": "alipay"}
                                acct_id = acct_map.get(acct_name, acct_name)
                        transactions.append({
                            "id": _uid(),
                            "account_id": str(acct_id),
                            "type": str(record.get("type", "pnl")),
                            "date": str(record.get("date", "")),
                            "amount": float(record.get("amount", 0) or 0),
                            "currency": str(record.get("currency", "CNY")),
                            "notes": str(record.get("notes", "")),
                        })

                elif target_type == "deposit":
                    if "amount" in record and record["amount"] is not None:
                        acct_id = record.get("account_id", "deposit")
                        deposits.append({
                            "id": _uid(),
                            "account_id": str(acct_id),
                            "amount": float(record.get("amount", 0) or 0),
                            "rate": float(record.get("rate", 0) or 0),
                            "start_date": str(record.get("start_date", "")),
                            "maturity_date": str(record.get("maturity_date", "")),
                            "interest": float(record.get("interest", 0) or 0),
                        })

                elif target_type == "exchange_rate":
                    # Handled via cell_references above
                    pass

    wb.close()

    return {
        "accounts": accounts,
        "holdings": holdings,
        "transactions": transactions,
        "deposits": deposits,
        "rates": rates_data,
    }

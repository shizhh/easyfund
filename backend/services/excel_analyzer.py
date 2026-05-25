"""Excel metadata extraction and AI mapping analysis."""

import json
import logging
import re

import httpx
import openpyxl

from backend.schemas import ImportMappingSchema

logger = logging.getLogger(__name__)

# ── System prompt for AI analyzer ─────────────────────────────

ANALYZER_SYSTEM_PROMPT = """\
You are a data mapping assistant for EasyFund (墨账), a personal finance management application.
Your task is to analyze an Excel workbook's structure and produce a JSON mapping that describes
how each sheet and column should be imported into EasyFund's data model.

EasyFund data model:
- Account: {id, name, category (cash|investment|insurance|future_cash), institution, balances: [{currency, amount}], notes}
- Holding: {id, account_id, ticker, name, shares, avg_cost_price, current_price, currency, vested_shares, unvested_shares}
- Transaction: {id, account_id, type (deposit|withdraw|buy|sell|dividend|interest|pnl), date, amount, currency, quantity, price, pnl, notes}
- Deposit: {id, account_id, amount, rate, start_date, maturity_date, interest}
- ExchangeRate: {currency, rate_to_cny, date}

Rules:
1. Each sheet maps to one or more target types (Account, Holding, Transaction, Deposit, ExchangeRate).
2. Each column maps to a field on the target type, or is marked as "skip".
3. Category mapping: map raw cell values to one of: cash, investment, insurance, future_cash.
4. Currency detection: if a column contains mixed currencies, specify the parsing rule (regex or fixed).
5. If exchange rates are stored in specific cells (e.g., H2, I2), identify those cell references.
6. If a column contains embedded data (e.g., "12022HKD" or "剩下902股"), specify the extraction regex as a transform.
7. Return ONLY valid JSON matching the ImportMappingSchema format. No explanations outside JSON.

Transform format:
- "regex:PATTERN" — apply re.search(PATTERN, value), use transform_group to pick capture group
- "int" — cast to int
- "float" — cast to float
- "date:FORMAT" — parse date with strptime format
- "currency_code" — extract 3-letter uppercase currency code from text
"""


# ── Metadata Extraction ───────────────────────────────────────

def extract_workbook_metadata(content: bytes) -> dict:
    """Extract structural metadata from an Excel workbook for AI analysis."""
    import io

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    sheets_meta = []

    for ws in wb.worksheets:
        # Determine column count from first few rows
        max_col = 0
        for row in ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=False):
            for cell in row:
                if cell.value is not None and cell.column > max_col:
                    max_col = cell.column
        max_col = min(max_col, 30)  # cap at 30 columns

        # Extract headers from first row
        headers = []
        for col_idx in range(1, max_col + 1):
            val = ws.cell(row=1, column=col_idx).value
            headers.append(str(val) if val is not None else "")

        # Extract sample data rows (up to 10)
        sample_rows = []
        for row in ws.iter_rows(min_row=2, max_row=min(11, ws.max_row), values_only=False):
            row_data = []
            for col_idx in range(1, max_col + 1):
                cell = row[col_idx - 1] if col_idx - 1 < len(row) else None
                val = cell.value if cell else None
                row_data.append(val)
            sample_rows.append(row_data)

        # Detect notable standalone cells (cells with values outside the main data area,
        # typically in the header area, e.g., exchange rate references like H2, I2)
        notable_cells = {}
        for row_idx in range(1, min(4, ws.max_row + 1)):
            for col_idx in range(max_col + 1, min(max_col + 6, ws.max_column + 1)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    notable_cells[f"{col_letter}{row_idx}"] = val

        # Detect merged cell ranges
        merged_ranges = []
        for merged in ws.merged_cells.ranges:
            merged_ranges.append(str(merged))

        sheets_meta.append({
            "name": ws.title,
            "headers": headers,
            "sample_rows": sample_rows,
            "notable_cells": notable_cells,
            "merged_ranges": merged_ranges,
            "total_rows": ws.max_row,
            "total_columns": max_col,
        })

    wb.close()
    return {"sheets": sheets_meta}


# ── AI Mapping Call ───────────────────────────────────────────

async def call_ai_for_mapping(metadata: dict) -> ImportMappingSchema:
    """Call AI to analyze workbook metadata and return a mapping schema."""
    from backend.services.ai_service import load_ai_config

    config = load_ai_config()
    api_key = config.get("api_key", "")
    if not api_key:
        raise ValueError("AI 未配置，请先在 AI 助手设置中填写 API Key")

    base_url = config.get("base_url", "https://api.openai.com/v1").rstrip("/")
    model = config.get("model", "gpt-4o-mini")

    url = f"{base_url}/chat/completions"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }

    body = {
        "model": model,
        "messages": [
            {"role": "system", "content": ANALYZER_SYSTEM_PROMPT},
            {"role": "user", "content": json.dumps(metadata, ensure_ascii=False, default=str)},
        ],
        "temperature": 0.1,
        "response_format": {"type": "json_object"},
    }

    for attempt in range(2):
        try:
            async with httpx.AsyncClient(timeout=60.0) as client:
                resp = await client.post(url, json=body, headers=headers)
                resp.raise_for_status()
                data = resp.json()

            content = data["choices"][0]["message"]["content"]
            mapping_dict = json.loads(content)
            return ImportMappingSchema(**mapping_dict)

        except Exception as e:
            if attempt == 0:
                logger.warning("AI mapping attempt 1 failed: %s, retrying...", e)
                # On retry, add a stricter instruction
                body["messages"].append({
                    "role": "assistant",
                    "content": "I will return valid JSON matching the ImportMappingSchema.",
                })
                body["messages"].append({
                    "role": "user",
                    "content": "Please return ONLY the JSON mapping, ensuring it matches the ImportMappingSchema format exactly.",
                })
            else:
                logger.error("AI mapping failed after 2 attempts: %s", e)
                raise ValueError(f"AI 分析失败，无法生成有效的映射配置: {e}") from e
